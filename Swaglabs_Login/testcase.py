from selenium import webdriver
import openpyxl
import time
driver = webdriver.Chrome("C:/Users\Gilta\Downloads\chromedriver_win32\chromedriver.exe")
wb= openpyxl.load_workbook("C:/Users\Gilta\PycharmProjects\Excel\credentials.xlsx")
sheet=wb.active
#data= sheet['B2'].value
rows=sheet.max_row
cols=sheet.max_column
driver.get("https://www.saucedemo.com/")
for i in range(2,rows+1):
    driver = webdriver.Chrome("C:/Users\Gilta\Downloads\chromedriver_win32\chromedriver.exe")
    driver.get("https://www.saucedemo.com/")
    username = sheet.cell(row=i, column=1).value
    print("username : ", username)
    password = sheet.cell(row=i, column=2).value
    print("password :", password)
    driver.find_element_by_id("user-name").send_keys(username)
    driver.find_element_by_id("password").send_keys(password)
    driver.find_element_by_id("login-button").click()
    time.sleep(5)
    driver.close()




'''
driver.get("https://www.saucedemo.com/")
driver.close()

'''